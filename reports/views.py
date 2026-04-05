"""Reports & Analytics views."""

import csv
import io
from datetime import date, datetime, timedelta
from decimal import Decimal

from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models import Avg, Count, Max, Min, Q, Sum
from django.http import HttpResponse
from django.shortcuts import render
from django.utils import timezone
from django.views.generic import View
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from orders.models import Order, OrderStatus
from references.models import Client
from transport.models import Driver, TripLog, Vehicle, VehicleStatus


# ─── Утилиты ─────────────────────────────────────────────────────

def _get_date_range(request):
    """Получить диапазон дат из GET-параметров."""
    date_from = request.GET.get("date_from")
    date_to = request.GET.get("date_to")

    if date_from:
        date_from = datetime.strptime(date_from, "%Y-%m-%d").date()
    else:
        date_from = timezone.now().date() - timedelta(days=30)

    if date_to:
        date_to = datetime.strptime(date_to, "%Y-%m-%d").date()
    else:
        date_to = timezone.now().date()

    return date_from, date_to


def _export_to_excel(filename, headers, rows, title=None):
    """Создать HttpResponse с Excel-файлом."""
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    wb = Workbook()
    ws = wb.active
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    # Заголовок отчёта
    if title:
        ws.merge_cells("A1:G1")
        cell = ws["A1"]
        cell.value = title
        cell.font = Font(bold=True, size=14)
        cell.alignment = Alignment(horizontal="center")
        ws.append([])

    # Шапка таблицы
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    ws.append(headers)
    for cell in ws[ws.max_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Данные
    for row in rows:
        ws.append([_format_cell(v) for v in row])

    # Автоширина колонок (пропускаем MergedCell)
    from openpyxl.utils import get_column_letter
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for row_num in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_num, column=col_idx)
            try:
                if cell.value:
                    val_len = len(str(cell.value))
                    if val_len > max_len:
                        max_len = val_len
            except (AttributeError, TypeError):
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 40)

    wb.save(response)
    return response


def _export_to_csv(headers, rows, filename="report.csv"):
    """Создать HttpResponse с CSV-файлом."""
    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    writer = csv.writer(response)
    # BOM для корректного открытия в Excel
    response.write("\ufeff")
    writer.writerow(headers)
    for row in rows:
        writer.writerow([str(v) for v in row])

    return response


def _format_cell(value):
    """Форматировать значение для ячейки."""
    if isinstance(value, datetime):
        return value.strftime("%d.%m.%Y %H:%M")
    if isinstance(value, date):
        return value.strftime("%d.%m.%Y")
    if isinstance(value, Decimal):
        return round(value, 2)
    if value is None:
        return ""
    return value


def _handle_export(request, excel_response, csv_content):
    """Выбрать формат экспорта."""
    fmt = request.GET.get("export")
    if fmt == "csv":
        return csv_content
    return excel_response


# ─── Главная страница отчётов ────────────────────────────────────


def reports_index(request):
    """Главная страница отчётов — карточки разделов."""
    now = timezone.now()
    today = now.date()
    month_start = today.replace(day=1)

    ctx = {
        # Общие метрики
        "total_orders": Order.objects.count(),
        "total_clients": Client.objects.count(),
        "total_vehicles": Vehicle.objects.count(),
        "total_drivers": Driver.objects.count(),
        # За этот месяц
        "orders_this_month": Order.objects.filter(created_at__date__gte=month_start).count(),
        "completed_this_month": Order.objects.filter(
            status=OrderStatus.COMPLETED, updated_at__date__gte=month_start
        ).count(),
        # Транспорт
        "vehicles_on_line": Vehicle.objects.filter(status=VehicleStatus.ON_LINE).count(),
        "vehicles_in_repair": Vehicle.objects.filter(status=VehicleStatus.REPAIR).count(),
        # Последние рейсы
        "recent_trips": TripLog.objects.select_related("driver", "vehicle").order_by(
            "-departure_date"
        )[:5],
    }
    return render(request, "reports/index.html", ctx)


# ─── Сводный дашборд ─────────────────────────────────────────────


def dashboard(request):
    """Сводный дашборд с графиками и метриками."""
    now = timezone.now()
    today = now.date()
    week_ago = today - timedelta(days=7)
    month_ago = today - timedelta(days=30)

    # KPI
    active_orders = Order.objects.filter(
        status__in=[
            OrderStatus.CONFIRMED,
            OrderStatus.ASSIGNED,
            OrderStatus.IN_TRANSIT,
            OrderStatus.LOADED,
            OrderStatus.UNLOADED,
        ]
    ).count()

    delayed_orders = Order.objects.filter(
        status__in=[OrderStatus.IN_TRANSIT, OrderStatus.ASSIGNED, OrderStatus.LOADED],
        required_delivery_datetime__lt=now,
    ).count()

    vehicles_on_line = Vehicle.objects.filter(status=VehicleStatus.ON_LINE).count()
    completed_today = Order.objects.filter(
        status=OrderStatus.COMPLETED, updated_at__date=today
    ).count()

    # Статистика за месяц
    month_orders = Order.objects.filter(created_at__date__gte=month_ago)
    month_completed = month_orders.filter(status=OrderStatus.COMPLETED).count()
    month_cancelled = month_orders.filter(status=OrderStatus.CANCELLED).count()
    month_weight = month_orders.aggregate(total=Sum("cargo_weight_kg"))["total"] or 0
    month_revenue_orders = month_orders.count()  # можно привязать к деньгам

    # Статистика по ТС
    vehicle_stats = {
        "total": Vehicle.objects.count(),
        "ok": Vehicle.objects.filter(status=VehicleStatus.OK).count(),
        "on_line": vehicles_on_line,
        "repair": Vehicle.objects.filter(status=VehicleStatus.REPAIR).count(),
        "write_off": Vehicle.objects.filter(status=VehicleStatus.WRITE_OFF).count(),
    }

    # Заказы по статусам (для диаграммы)
    status_data = []
    for code, label in OrderStatus.choices:
        cnt = Order.objects.filter(status=code).count()
        if cnt > 0:
            status_data.append({"label": label, "count": cnt})

    # Заказы за последние 7 дней (для линейного графика)
    daily_orders = []
    for i in range(6, -1, -1):
        d = today - timedelta(days=i)
        cnt = Order.objects.filter(created_at__date=d).count()
        daily_orders.append({"date": d.strftime("%d.%m"), "count": cnt})

    # Топ-5 клиентов по количеству заказов
    top_clients = (
        Client.objects.annotate(order_count=Count("orders"))
        .filter(order_count__gt=0)
        .order_by("-order_count")[:5]
    )

    # Топ-5 водителей по количеству заказов
    top_drivers = (
        Driver.objects.annotate(order_count=Count("orders"))
        .filter(order_count__gt=0, is_active=True)
        .order_by("-order_count")[:5]
    )

    # Журнал рейсов за неделю
    week_trips = TripLog.objects.filter(departure_date__gte=week_ago).select_related(
        "driver", "vehicle"
    )
    week_mileage = week_trips.aggregate(total=Sum("mileage"))["total"] or 0
    week_fuel = week_trips.aggregate(total=Sum("fuel_consumed"))["total"] or 0

    ctx = {
        "kpi": {
            "active_orders": active_orders,
            "delayed_orders": delayed_orders,
            "vehicles_on_line": vehicles_on_line,
            "completed_today": completed_today,
        },
        "month": {
            "orders": month_orders.count(),
            "completed": month_completed,
            "cancelled": month_cancelled,
            "weight_kg": round(month_weight, 0),
        },
        "vehicle_stats": vehicle_stats,
        "status_data": status_data,
        "daily_orders": daily_orders,
        "top_clients": top_clients,
        "top_drivers": top_drivers,
        "week_trips": week_trips[:10],
        "week_mileage": week_mileage,
        "week_fuel": round(week_fuel, 1),
    }
    return render(request, "reports/dashboard.html", ctx)


# ─── Отчёт по заказам ────────────────────────────────────────────


def orders_report(request):
    """Отчёт по заказам: фильтрация по датам, клиенту, статусу."""
    date_from, date_to = _get_date_range(request)

    # Базовый queryset
    qs = Order.objects.select_related("client", "vehicle", "driver", "cargo_type").filter(
        created_at__date__range=(date_from, date_to)
    )

    # Фильтр по клиенту
    client_id = request.GET.get("client_id")
    if client_id:
        qs = qs.filter(client_id=client_id)

    # Фильтр по статусу
    status = request.GET.get("status")
    if status:
        qs = qs.filter(status=status)

    # Фильтр по водителю
    driver_id = request.GET.get("driver_id")
    if driver_id:
        qs = qs.filter(driver_id=driver_id)

    orders = qs.order_by("-created_at")

    # Агрегаты
    total = orders.count()
    total_weight = orders.aggregate(total=Sum("cargo_weight_kg"))["total"] or 0
    total_volume = orders.aggregate(total=Sum("cargo_volume_m3"))["total"] or 0
    completed = orders.filter(status=OrderStatus.COMPLETED).count()
    cancelled = orders.filter(status=OrderStatus.CANCELLED).count()

    # Распределение по статусам
    by_status = []
    for code, label in OrderStatus.choices:
        cnt = orders.filter(status=code).count()
        if cnt > 0:
            by_status.append({"label": label, "count": cnt})

    # Распределение по клиентам
    by_client = (
        orders.values("client__name")
        .annotate(count=Count("id"), weight=Sum("cargo_weight_kg"))
        .order_by("-count")[:15]
    )

    ctx = {
        "orders": orders[:200],  # лимит для отображения
        "total": total,
        "total_weight": round(total_weight, 0),
        "total_volume": round(total_volume, 1),
        "completed": completed,
        "cancelled": cancelled,
        "by_status": by_status,
        "by_client": by_client,
        "date_from": date_from,
        "date_to": date_to,
        "clients": Client.objects.filter(orders__isnull=False).distinct()[:50],
        "current_client": client_id or "",
        "current_status": status or "",
    }

    # Экспорт
    if request.GET.get("export") in ("excel", "csv"):
        return _export_orders(request, orders)

    return render(request, "reports/orders_report.html", ctx)


def _export_orders(request, orders):
    """Экспорт отчёта по заказам."""
    headers = ["№ заказа", "Клиент", "Груз", "Вес (кг)", "Объём (м³)", "Статус", "ТС", "Водитель", "Создан"]
    rows = []
    for o in orders:
        rows.append([
            o.order_number,
            o.client.name,
            o.cargo_name,
            o.cargo_weight_kg,
            o.cargo_volume_m3,
            o.get_status_display(),
            o.vehicle.plate_number if o.vehicle else "—",
            o.driver.full_name if o.driver else "—",
            o.created_at,
        ])

    excel_resp = _export_to_excel("orders_report.xlsx", headers, rows, title="Отчёт по заказам")
    csv_resp = _export_to_csv(headers, rows, "orders_report.csv")
    return _handle_export(request, excel_resp, csv_resp)


# ─── Отчёт по ТС ─────────────────────────────────────────────────


def vehicles_report(request):
    """Отчёт по эффективности транспортных средств."""
    date_from, date_to = _get_date_range(request)

    vehicles = Vehicle.objects.all().select_related()

    # Статистика по каждому ТС
    vehicle_data = []
    for v in vehicles:
        trips = TripLog.objects.filter(vehicle=v, departure_date__date__range=(date_from, date_to))
        orders = Order.objects.filter(vehicle=v, created_at__date__range=(date_from, date_to))

        total_mileage = trips.aggregate(total=Sum("mileage"))["total"] or 0
        total_fuel = trips.aggregate(total=Sum("fuel_consumed"))["total"] or 0
        trip_count = trips.count()
        order_count = orders.count()

        # Средняя загрузка
        avg_weight = orders.aggregate(avg=Avg("cargo_weight_kg"))["avg"]
        avg_volume = orders.aggregate(avg=Avg("cargo_volume_m3"))["avg"]

        vehicle_data.append({
            "vehicle": v,
            "status": v.get_status_display(),
            "trip_count": trip_count,
            "order_count": order_count,
            "total_mileage": total_mileage,
            "total_fuel": round(total_fuel, 1),
            "avg_consumption": round(total_fuel / (total_mileage / 100), 1) if total_mileage > 0 else 0,
            "avg_weight": round(avg_weight, 0) if avg_weight else 0,
            "avg_volume": round(avg_volume, 1) if avg_volume else 0,
        })

    # Фильтр по статусу
    status_filter = request.GET.get("status")
    if status_filter:
        vehicle_data = [v for v in vehicle_data if v["vehicle"].status == status_filter]

    # Суммарные показатели
    total_trips = sum(v["trip_count"] for v in vehicle_data)
    total_mileage = sum(v["total_mileage"] for v in vehicle_data)
    total_fuel = sum(v["total_fuel"] for v in vehicle_data)

    # Для диаграммы
    status_counts = {}
    for v in vehicles:
        s = v.get_status_display()
        status_counts[s] = status_counts.get(s, 0) + 1

    ctx = {
        "vehicle_data": vehicle_data,
        "total_trips": total_trips,
        "total_mileage": total_mileage,
        "total_fuel": round(total_fuel, 1),
        "date_from": date_from,
        "date_to": date_to,
        "status_counts": status_counts,
        "current_status": status_filter or "",
    }

    if request.GET.get("export") in ("excel", "csv"):
        return _export_vehicles(request, vehicle_data)

    return render(request, "reports/vehicles_report.html", ctx)


def _export_vehicles(request, vehicle_data):
    """Экспорт отчёта по ТС."""
    headers = [
        "ТС", "Тип", "Статус", "Рейсов", "Заказов",
        "Пробег (км)", "Топливо (л)", "Ср. расход (л/100км)",
        "Ср. вес (кг)", "Ср. объём (м³)",
    ]
    rows = []
    for v in vehicle_data:
        rows.append([
            f"{v['vehicle'].plate_number} ({v['vehicle'].brand} {v['vehicle'].model})",
            v["vehicle"].get_vehicle_type_display(),
            v["status"],
            v["trip_count"],
            v["order_count"],
            v["total_mileage"],
            v["total_fuel"],
            v["avg_consumption"],
            v["avg_weight"],
            v["avg_volume"],
        ])

    excel_resp = _export_to_excel("vehicles_report.xlsx", headers, rows, title="Отчёт по ТС")
    csv_resp = _export_to_csv(headers, rows, "vehicles_report.csv")
    return _handle_export(request, excel_resp, csv_resp)


# ─── Отчёт по водителям ──────────────────────────────────────────


def drivers_report(request):
    """Отчёт по эффективности водителей."""
    date_from, date_to = _get_date_range(request)

    drivers = Driver.objects.filter(is_active=True)

    driver_data = []
    for d in drivers:
        trips = TripLog.objects.filter(driver=d, departure_date__date__range=(date_from, date_to))
        orders = Order.objects.filter(driver=d, created_at__date__range=(date_from, date_to))

        total_mileage = trips.aggregate(total=Sum("mileage"))["total"] or 0
        total_fuel = trips.aggregate(total=Sum("fuel_consumed"))["total"] or 0
        trip_count = trips.count()
        order_count = orders.count()

        # Среднее время рейса
        completed_trips = trips.filter(arrival_date__isnull=False)
        avg_trip_hours = None
        if completed_trips.exists():
            total_hours = 0
            count = 0
            for t in completed_trips:
                if t.arrival_date and t.departure_date:
                    delta = t.arrival_date - t.departure_date
                    total_hours += delta.total_seconds() / 3600
                    count += 1
            if count > 0:
                avg_trip_hours = round(total_hours / count, 1)

        driver_data.append({
            "driver": d,
            "trip_count": trip_count,
            "order_count": order_count,
            "total_mileage": total_mileage,
            "total_fuel": round(total_fuel, 1),
            "avg_trip_hours": avg_trip_hours,
            "orders_completed": orders.filter(status=OrderStatus.COMPLETED).count(),
            "orders_cancelled": orders.filter(status=OrderStatus.CANCELLED).count(),
        })

    # Сортировка по количеству рейсов
    sort_by = request.GET.get("sort", "trip_count")
    reverse = request.GET.get("reverse", "true") == "true"
    if sort_by in ("trip_count", "order_count", "total_mileage", "total_fuel"):
        driver_data.sort(key=lambda x: x[sort_by], reverse=reverse)

    # Топ-10
    top_drivers = sorted(driver_data, key=lambda x: x["trip_count"], reverse=True)[:10]

    ctx = {
        "driver_data": driver_data,
        "top_drivers": top_drivers,
        "date_from": date_from,
        "date_to": date_to,
        "current_sort": sort_by,
        "current_reverse": reverse,
    }

    if request.GET.get("export") in ("excel", "csv"):
        return _export_drivers(request, driver_data)

    return render(request, "reports/drivers_report.html", ctx)


def _export_drivers(request, driver_data):
    """Экспорт отчёта по водителям."""
    headers = [
        "Водитель", "Телефон", "Категория прав",
        "Рейсов", "Заказов", "Завершено", "Отменено",
        "Пробег (км)", "Топливо (л)", "Ср. время рейса (ч)",
    ]
    rows = []
    for d in driver_data:
        rows.append([
            d["driver"].full_name,
            d["driver"].phone or "—",
            d["driver"].license_category,
            d["trip_count"],
            d["order_count"],
            d["orders_completed"],
            d["orders_cancelled"],
            d["total_mileage"],
            d["total_fuel"],
            d["avg_trip_hours"] or "—",
        ])

    excel_resp = _export_to_excel("drivers_report.xlsx", headers, rows, title="Отчёт по водителям")
    csv_resp = _export_to_csv(headers, rows, "drivers_report.csv")
    return _handle_export(request, excel_resp, csv_resp)
